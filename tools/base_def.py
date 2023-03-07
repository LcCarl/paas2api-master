# -*- coding: utf-8 -*-
# coding: utf-8
# @Time    : 2020/7/6 20:10
# @Author  : Jeico
# @File    : base_def.py

import time
import datetime
import random
import importlib
import hashlib
import pandas as pd
from openpyxl import load_workbook
from jsonpath_rw import parse
from tools.my_log import MyLog
from data_init.common_data import CommonData
from conf.environment_config import EnvironmentConfig

my_logger = MyLog()


class BaseDef:

    @staticmethod
    def write_back(file_name, sheet_name, row, col, result):
        """——————数据写回表格————————
            参数：
                file_name：文件名   str； sheet_name：表名  str；row：行  int ；
                col：列    int  ； result：结果值   str
            ——————————————————————————————————————————"""
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, col).value = result
        wb.save(file_name)

    @staticmethod
    def read_excel(file_name, sheet_name, row, col):
        """——————读取表格其中一格————————
        参数：
            file_name:文件名  str； sheet_name：表名  str；
            row：实际行数-2    int； col：实际列数-1    int；
        返回： 表格内容    str
        ——————————————————————————————————————————"""
        excel_data = pd.read_excel(file_name, sheet_name=sheet_name).iloc[row, col]
        return excel_data

    @staticmethod
    def split_data(data):
        """——————分割xxx>xxx>xxx格式的字符串，并分别传出————————
            参数：
                data：格式参照上面   str;
            返回： 分割后的三段字符串     str
            ——————————————————————————————————————————"""
        sheet_name = data.split(">")[0]
        case_id = data.split(">")[1]
        rule_data = data.split(">")[2]
        return sheet_name, case_id, rule_data

    @staticmethod
    def get_depend_data(res_data, res_key):
        """——————根据res_data，查询关键字的值————————
            参数：
                res_data:json数据；  str  ；
                res_key：查询的关键字路径  str；
            返回：  关键字的值   str
            ——————————————————————————————————————————"""
        json_exe = parse(res_key)
        res_data = eval(res_data)
        male = json_exe.find(res_data)
        for math in male:
            return math.value

    @staticmethod
    def replace_data(res_data, res_key, value):
        """——————根据res_key，替换掉res_data里面对应的参数值————————
            参数：
                res_data:接口的传参  dict  ；
                res_key：需要替换值的关键参数路径  str； eg:"platformProjects[*].platformId"
                value：被替换的值 str；
            返回：  替换后的参数   str
            ——————————————————————————————————————————"""
        dict_data = res_data
        key_list = res_key.split(".")
        key_length = len(key_list)
        for position, i in enumerate(key_list):
            if "[*]" in i:
                e = i.split("[")[0]
                dict_data = dict_data[e][0]
            else:
                if position + 1 == key_length:
                    dict_data[i] = value
                else:
                    dict_data = dict_data[i]
        return res_data

    @staticmethod
    def time_stamp():
        """ ——————根据时间戳生成随机数——————
        返回：  时间戳+3位随机数   int  13位
        ——————————————————————————————————————————"""

        def t2():
            return int(time.time()) * 1000 + random.randint(100, 999)

        return t2

    def get_email(self):
        """ ——————根据时间戳生成邮箱——————
        返回：  邮箱号  str
        ——————————————————————————————————————————"""
        my_timestamp = self.time_stamp()()
        return str(my_timestamp) + "@qq.com"

    def get_phone(self):
        """ ——————根据时间戳生成邮箱——————
        返回：  手机号  str
        ——————————————————————————————————————————"""
        my_timestamp = self.time_stamp()()
        return "130" + str(my_timestamp)[5:]

    @staticmethod
    def do_time_report():
        """——————将接口测试耗时进行处理——————"""
        time_list = getattr(CommonData, "TimeList")  # 从公共数据读取数据
        api_dic = getattr(CommonData, "TimeConsuming")
        if time_list and api_dic:
            max_time = max(time_list)
            min_time = min(time_list)
            sum_time = sum(time_list)
            num = len(time_list)
            mid_time = sum_time / num

            max_time_api = []
            for m, n in api_dic.items():
                if n == max_time:
                    max_time_api.append(m)
            result = "本次测试，共测试{}个接口，共耗时{}秒，平均耗时{}秒，单个接口，耗时最短{}秒，最长{}秒，接口为{}。" \
                .format(str(num), str(sum_time), str(mid_time), str(min_time), str(max_time), str(max_time_api))
        else:
            result = "耗时计算失败"
        return result

    @staticmethod
    def get_import(init_model_path):
        """——————根据文件名，获取此文件中，类名包含‘Init’的类名——————
        参数：
            init_model_path ：init文件文件名  str ；
        返回： 模块名      class ;
        ——————————————————————————————————————————"""
        full_path = EnvironmentConfig().test_init + init_model_path
        init_params = importlib.import_module(full_path)
        init_model_list = dir(init_params)
        goal_mode_list = [goal_mode for goal_mode in init_model_list if "Init" in goal_mode]
        init_model = getattr(init_params, goal_mode_list[0])
        return init_model

    @staticmethod
    def random_number():
        """——————根据最小最大值，获取随机数。取小不取大——————"""

        def n1(start, end):
            return random.randint(start, end)

        return n1

    @staticmethod
    def md5ing(obj):
        """——————MD5加密参数——————
        参数：
            obj ：待加密的内容  str ；
        返回： 加密文本    str ;
        ——————————————————————————————————————————"""
        try:
            new_obj = obj.encode()
            m = hashlib.md5()
            m.update(new_obj)
            return m.hexdigest().upper()
        except Exception as e:
            raise e

    @staticmethod
    def query_time(span="day"):
        """ ——————————————————————————————————————
        返回：  当前时间，过去时间，未来时间  str
        ——————————————————————————————————————————"""
        now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if span == "day":
            past_time = (datetime.datetime.now() + datetime.timedelta(days=-1)).strftime("%Y-%m-%d %H:%M:%S")
            future_time = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
        else:
            past_time = (datetime.datetime.now() + datetime.timedelta(hours=-1)).strftime("%Y-%m-%d %H:%M:%S")
            future_time = (datetime.datetime.now() + datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
        return now_time, past_time, future_time
