# -*- coding: utf-8 -*-
# @Time    : 2020/11/20 14:41
# @Author  : Jeico
# @File    : processing_data.py

from openpyxl import load_workbook
from conf.environment_config import EnvironmentConfig
from tools.my_log import MyLog
from tools.do_config import DoConfig
from tools.regex_before_data import BeforeDataRegex
from tools.project_path import *
from tools.base_def import BaseDef

my_logger = MyLog()


class DataProcessing:

    @staticmethod
    def data_processing():
        """————————根据配置文件，从excel表中，获取并处理测试数据————————————"""
        test_data = []
        test_config_path = EnvironmentConfig.test_config_path
        mode_list = eval(DoConfig.read_config(test_config_path, "LIST", "mode_list"))

        for i in mode_list:
            mode = eval(DoConfig.read_config(test_config_path, "MODE", i))
            test_case_path = os.path.join(EnvironmentConfig.test_case_bag, i + ".xlsx")
            wb = load_workbook(test_case_path)

            for key in mode:
                sheet = wb[key]
                data_list = []
                if mode[key] == "all":
                    for item in range(2, sheet.max_row + 1):
                        data_list.append(item)
                else:
                    for item in mode[key]:
                        data_list.append(item + 1)

                for item in data_list:
                    sub_data = {"id": sheet.cell(item, 1).value, "sheet": sheet.cell(item, 2).value,
                                "module": sheet.cell(item, 3).value, "case_name": sheet.cell(item, 4).value,
                                "method": sheet.cell(item, 5).value, "api": sheet.cell(item, 6).value,
                                "data": sheet.cell(item, 7).value, "add_data": sheet.cell(item, 8).value,
                                "case_path": sheet.cell(item, 9).value, "init_path": sheet.cell(item, 10).value,
                                "code": sheet.cell(item, 12).value, "check_sql": sheet.cell(item, 15).value,
                                "returned_value": sheet.cell(item, 16).value, "expected": sheet.cell(item, 18).value,
                                "precondition": sheet.cell(item, 20).value, "key": sheet.cell(item, 21).value}
                    sub_data = BeforeDataRegex().before_data_regex(str(sub_data),
                                                                   BaseDef().get_import(sub_data["init_path"]))
                    test_data.append(eval(sub_data))

        return test_data
